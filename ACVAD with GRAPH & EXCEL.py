import cv2
import mediapipe as mp
import numpy as np
import matplotlib.pyplot as plt
import openpyxl  # Import the openpyxl library for Excel operations

mp_drawing = mp.solutions.drawing_utils
mp_pose = mp.solutions.pose

cap = cv2.VideoCapture(0)

desired_radius = int(input("Enter the desired radius for the arch: "))
spacing = 20  # Adjust as needed

angles_list = []  # List to store detected angles

plt.ion()  # Turn on interactive mode for Matplotlib

# Create initial empty plot
fig, ax = plt.subplots(figsize=(8, 6))
line, = ax.plot([], color='blue')
ax.set_title('Detected Angle Over Time')
ax.set_xlabel('Frame')
ax.set_ylabel('Angle (degrees)')
ax.grid()


def update_plot(angles):
    line.set_xdata(np.arange(len(angles)))
    line.set_ydata(angles)
    ax.relim()
    ax.autoscale_view()
    fig.canvas.draw()
    fig.canvas.flush_events()


def calculate_angle(a, b, c):
    a = np.array(a)  # First
    b = np.array(b)  # Mid
    c = np.array(c)  # End

    radians = np.arctan2(c[1] - b[1], c[0] - b[0]) - np.arctan2(a[1] - b[1], a[0] - b[0])
    angle = np.abs(radians * 180.0 / np.pi)

    rounded_angle = round(angle, 4)

    if angle > 180.0:
        angle = 360 - angle

    return rounded_angle


def draw_filled_parallel_arches(frame, center, radius, spacing, thickness=5, angle=0):
    """Draws two parallel vertical arches with connected ends and a filled spacing area."""
    start_angle = 0
    end_angle = 179  # Use full angle for filled area

    # Check the angle and change color accordingly
    if 80 <= angle <= 100:
        color = (0, 165, 255)  # Orange color for angle between 80 and 100
    elif angle < 80:
        color = (0, 0, 255)  # RED color for angle < 80
    else:
        color = (0, 255, 0)  # Green color for angle >= 100

        # Create points for the combined shape, ensuring overlap
    left_points = cv2.ellipse2Poly(
        (center[0] - spacing // 2, center[1]), (radius, radius // 2), 90, start_angle, end_angle, 1
    )
    right_points = cv2.ellipse2Poly(
        (center[0] + spacing // 2, center[1]), (radius, radius // 2), 90, start_angle, end_angle, 1
    )

    combined_points = np.concatenate((left_points, right_points[::-1]))

    # Draw as open polylines to avoid connecting lines
    cv2.polylines(frame, [combined_points], False, color, thickness)

    # Fill the combined shape
    cv2.fillConvexPoly(frame, combined_points, color)


with mp_pose.Pose(min_detection_confidence=0.5, min_tracking_confidence=0.5) as pose:
    while cap.isOpened():
        ret, frame = cap.read()

        frame_height, frame_width, _ = frame.shape

        # Define arch parameters for left corner placement
        center_x = desired_radius  # Left edge of the frame plus radius
        center_y = frame_height // 2  # Vertically centered
        center = (center_x, center_y)

        image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        image.flags.writeable = False

        results = pose.process(image)

        image.flags.writeable = True
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

        angle = 0  # Default angle

        try:
            landmarks = results.pose_landmarks.landmark

            # get coordinates
            hip = [landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].x,
                   landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y]
            shoulder = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                        landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
            elbow = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                     landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]

            angle = calculate_angle(hip, shoulder, elbow)

            angle_text = f"Angle: {angle:.2f} deg"
            cv2.putText(image, angle_text, (frame_width - 300, 75),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2, cv2.LINE_AA)
            cv2.putText(image, str(angle),
                        tuple(np.multiply(shoulder, [640, 480]).astype(int)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 2, cv2.LINE_AA)

        except:
            pass

        angles_list.append(angle)  # Append detected angle to the list
        # Create or open an Excel workbook
        workbook = openpyxl.Workbook()
        # Load the existing Excel workbook
        workbook = openpyxl.load_workbook("angles.xlsx")  # Load the existing file
        sheet = workbook.active  # Select the active sheet

        # Get the next available row to append data
        next_row = sheet.max_row + 1  # Start writing from the first empty row

        # Write the new angle value to the next available row
        sheet.cell(row=next_row, column=1).value = angle

        # Save the updated workbook
        workbook.save("angles.xlsx")

        # Drawing the arch with color based on angle
        draw_filled_parallel_arches(image, center, desired_radius, spacing, angle=angle)

        mp_drawing.draw_landmarks(image, results.pose_landmarks, mp_pose.POSE_CONNECTIONS)

        cv2.imshow('Mediapipe feed', image)

        update_plot(angles_list)  # Update the Matplotlib plot

        if cv2.waitKey(10) & 0xFF == ord('q'):
            break

cap.release()
cv2.destroyAllWindows()
cv2.waitKey(1)  # Ensure Matplotlib plot remains visible
